#!/usr/bin/env python3
"""
DCF Valuation Model — config-driven
=====================================
Build a formatted Excel DCF for any company by supplying a YAML config file.

Usage:
    python dcf_model.py <config.yaml>      # e.g. cmg_config.yaml
    python dcf_model.py                    # falls back to cmg_config.yaml

Output:
    {TICKER}_DCF_Model.xlsx  (written next to the config file)

Dependencies:
    pip install openpyxl pyyaml
"""

import sys
from pathlib import Path

import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ═══════════════════════════════════════════════════════════════════════════════
#  CONFIG LOADER
# ═══════════════════════════════════════════════════════════════════════════════

def load_config(path: str) -> dict:
    with open(path) as f:
        return yaml.safe_load(f)


# ═══════════════════════════════════════════════════════════════════════════════
#  STATIC LAYOUT  (row map; column offsets derived from config at runtime)
# ═══════════════════════════════════════════════════════════════════════════════

COL_LBL = 1   # Column A — row labels
COL_H0  = 2   # Column B — first historical year

R = {
    # ── Summary ──────────────────────────────────────────────────────────────
    "TITLE":    1,
    "SUM1":     3,
    "SUM2":     4,
    # ── Assumptions ──────────────────────────────────────────────────────────
    "AH":       6,
    "VALASSUMP":7,
    "WACC":     8,
    "TGR":      9,
    # ── Income Statement ─────────────────────────────────────────────────────
    "IS_H":    11,
    "REV":     12,
    "REVG":    13,
    "EBIT":    14,
    "EBITM":   15,
    "TAX":     16,
    "TAXR":    17,
    # ── Cash Flow Items ───────────────────────────────────────────────────────
    "CF_H":    19,
    "DA":      20,
    "DAPCT":   21,
    "CAPEX":   22,
    "CXPCT":   23,
    "NWC":     24,
    "NWPCT":   25,
    # ── DCF Analysis ─────────────────────────────────────────────────────────
    "DCF_H":   27,
    "DREV":    28,
    "DREVG":   29,
    "DEBIT":   30,
    "DEBITM":  31,
    "DTAX":    32,
    "DTAXR":   33,
    "EBIAT":   35,
    "DDA":     36,
    "DDAPCT":  37,
    "DCAPEX":  38,
    "DCXPCT":  39,
    "DNWC":    40,
    "DNWPCT":  41,
    "UFCF":    43,
    "PVFCF":   44,
    # ── Valuation Bridge ─────────────────────────────────────────────────────
    "TV":      46,
    "PVTV":    47,
    "EV":      49,
    "CASH":    50,
    "DEBTROW": 51,
    "EQV":     52,
    "SHARES":  53,
    "SP":      54,
}


# ═══════════════════════════════════════════════════════════════════════════════
#  STYLE CONSTANTS
# ═══════════════════════════════════════════════════════════════════════════════

CLR_BLUE  = "0000FF"   # Hardcoded inputs
CLR_BLACK = "000000"   # Calculated formulas
CLR_WHITE = "FFFFFF"
CLR_GRAY  = "808080"
HDR_BG    = "1F3864"   # Dark navy — section headers
ASSUMP_BG = "D6E4F0"   # Light blue — assumption cells
HIST_BG   = "F2F2F2"   # Light grey — historical data
YELLOW    = "FFFF00"   # Key output highlight

FMT_INT   = '#,##0_);(#,##0);"-"'   # Whole-number thousands
FMT_PCT   = '0%;(0%);"-"'           # Percentage, 0 decimal places
FMT_PCT1  = '0.0%;(0.0%);"-"'       # Percentage, 1 decimal place (CapEx %)
FMT_PRICE = '$#,##0.00'             # Dollar price


# ═══════════════════════════════════════════════════════════════════════════════
#  STYLE HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def _font(bold=False, color=CLR_BLACK, size=9, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _align(h="right", v="center"):
    return Alignment(horizontal=h, vertical=v)

def _c(row, col):
    """Relative cell reference — e.g. G12."""
    return f"{get_column_letter(col)}{row}"

def _ca(row, col):
    """Absolute cell reference — e.g. $G$12."""
    return f"${get_column_letter(col)}${row}"

def _set(ws, row, col, value,
         bold=False, color=CLR_BLACK, fill_color=None,
         align_h="right", fmt=None, italic=False, size=9):
    cell = ws.cell(row=row, column=col)
    cell.value     = value
    cell.font      = _font(bold=bold, color=color, size=size, italic=italic)
    cell.alignment = _align(align_h)
    if fill_color:
        cell.fill = _fill(fill_color)
    if fmt:
        cell.number_format = fmt


# ═══════════════════════════════════════════════════════════════════════════════
#  SECTION WRITERS
# ═══════════════════════════════════════════════════════════════════════════════

def _write_title(ws, cfg):
    _set(ws, R["TITLE"], COL_LBL,
         f"{cfg['company_name']} — DCF Valuation Model",
         bold=True, size=13, align_h="left")


def _write_summary(ws, cfg, col_p_last, col_out):
    r1, r2 = R["SUM1"], R["SUM2"]
    sp_ref      = _ca(R["SP"],  col_out)
    implied_ref = _ca(r1, COL_H0 + 3)
    today_ref   = _ca(r1, COL_H0 + 6)

    _set(ws, r1, COL_LBL,       "Ticker:",              bold=True, align_h="left")
    _set(ws, r1, COL_H0,        cfg["ticker"],           bold=True, color=CLR_BLUE,
         fill_color=YELLOW, align_h="center")
    _set(ws, r1, COL_H0 + 2,   "Implied Share Price:",  bold=True, align_h="right")
    _set(ws, r1, COL_H0 + 3,   f"={sp_ref}",
         color=CLR_BLACK, fmt=FMT_PRICE, align_h="center")
    _set(ws, r1, COL_H0 + 5,   "Today's Share Price:",  bold=True, align_h="right")
    _set(ws, r1, COL_H0 + 6,   cfg["today_price"],
         color=CLR_BLUE, fmt=FMT_PRICE, align_h="center")
    _set(ws, r1, COL_H0 + 7,   "Upside / (Downside):", bold=True, align_h="right")
    _set(ws, r1, COL_H0 + 8,
         f"=({implied_ref}-{today_ref})/{today_ref}",
         color=CLR_BLACK, fmt=FMT_PCT, align_h="center")

    _set(ws, r2, COL_LBL, "Date:",            bold=True, align_h="left")
    _set(ws, r2, COL_H0,  cfg["model_date"],  color=CLR_BLUE, align_h="center")
    _set(ws, r2, col_p_last, "All numbers in thousands",
         italic=True, size=8, align_h="right", color=CLR_GRAY)


def _write_assumptions(ws, cfg):
    _set(ws, R["AH"],        COL_LBL, "Assumptions",
         bold=True, fill_color=HDR_BG, color=CLR_WHITE, align_h="left", size=10)
    _set(ws, R["VALASSUMP"], COL_LBL, "Valuation Assumptions",
         bold=True, align_h="left")
    _set(ws, R["WACC"],      COL_LBL, "WACC", bold=True, align_h="left")
    _set(ws, R["WACC"],      COL_H0,  cfg["wacc"],
         color=CLR_BLUE, fmt=FMT_PCT1, fill_color=ASSUMP_BG, align_h="center")
    _set(ws, R["TGR"],       COL_LBL, "TGR",  bold=True, align_h="left")
    _set(ws, R["TGR"],       COL_H0,  cfg["tgr"],
         color=CLR_BLUE, fmt=FMT_PCT1, fill_color=ASSUMP_BG, align_h="center")


def _section_header(ws, row_key, label, hist_years, proj_years, col_p0):
    _set(ws, R[row_key], COL_LBL, label,
         bold=True, fill_color=HDR_BG, color=CLR_WHITE, align_h="left")
    for i, yr in enumerate(hist_years):
        _set(ws, R[row_key], COL_H0 + i, str(yr),
             bold=True, fill_color=HDR_BG, color=CLR_WHITE, align_h="center")
    for i, yr in enumerate(proj_years):
        _set(ws, R[row_key], col_p0 + i, str(yr),
             bold=True, fill_color=HDR_BG, color=CLR_WHITE, align_h="center")


def _write_income_statement(ws, cfg, col_p0):
    hist = cfg["historical"]
    proj = cfg["projection"]
    n_hist = len(hist["years"])
    n_proj = len(proj["years"])

    _section_header(ws, "IS_H", "Income Statement",
                    hist["years"], proj["years"], col_p0)

    # ── Revenue ──────────────────────────────────────────────────────────────
    _set(ws, R["REV"], COL_LBL, "Revenue", bold=True, align_h="left")
    for i, v in enumerate(hist["revenue"]):
        _set(ws, R["REV"], COL_H0 + i, v,
             color=CLR_BLUE, fmt=FMT_INT, fill_color=HIST_BG)
    for i in range(n_proj):
        col  = col_p0 + i
        prev = _c(R["REV"],  col - 1)
        grw  = _c(R["REVG"], col)
        _set(ws, R["REV"], col, f"={prev}*(1+{grw})", color=CLR_BLACK, fmt=FMT_INT)

    # ── Revenue Growth % ─────────────────────────────────────────────────────
    _set(ws, R["REVG"], COL_LBL, "   % growth", italic=True, align_h="left")
    for i in range(1, n_hist):
        col = COL_H0 + i
        _set(ws, R["REVG"], col,
             f"=IFERROR(({_c(R['REV'],col)}-{_c(R['REV'],col-1)})/{_c(R['REV'],col-1)},0)",
             color=CLR_BLACK, fmt=FMT_PCT, fill_color=HIST_BG)
    for i, v in enumerate(proj["rev_growth"]):
        _set(ws, R["REVG"], col_p0 + i, v, color=CLR_BLUE, fmt=FMT_PCT)

    # ── EBIT ─────────────────────────────────────────────────────────────────
    _set(ws, R["EBIT"], COL_LBL, "EBIT", bold=True, align_h="left")
    for i, v in enumerate(hist["ebit"]):
        _set(ws, R["EBIT"], COL_H0 + i, v,
             color=CLR_BLUE, fmt=FMT_INT, fill_color=HIST_BG)
    for i in range(n_proj):
        col = col_p0 + i
        _set(ws, R["EBIT"], col,
             f"={_c(R['REV'],col)}*{_c(R['EBITM'],col)}",
             color=CLR_BLACK, fmt=FMT_INT)

    # ── EBIT Margin % ────────────────────────────────────────────────────────
    _set(ws, R["EBITM"], COL_LBL, "   % of sales", italic=True, align_h="left")
    for i in range(n_hist):
        col = COL_H0 + i
        _set(ws, R["EBITM"], col,
             f"=IFERROR({_c(R['EBIT'],col)}/{_c(R['REV'],col)},0)",
             color=CLR_BLACK, fmt=FMT_PCT, fill_color=HIST_BG)
    for i, v in enumerate(proj["ebit_margin"]):
        _set(ws, R["EBITM"], col_p0 + i, v, color=CLR_BLUE, fmt=FMT_PCT)

    # ── Taxes ────────────────────────────────────────────────────────────────
    _set(ws, R["TAX"], COL_LBL, "Taxes", bold=True, align_h="left")
    for i, v in enumerate(hist["taxes"]):
        _set(ws, R["TAX"], COL_H0 + i, v,
             color=CLR_BLUE, fmt=FMT_INT, fill_color=HIST_BG)
    for i in range(n_proj):
        col = col_p0 + i
        _set(ws, R["TAX"], col,
             f"={_c(R['EBIT'],col)}*{_c(R['TAXR'],col)}",
             color=CLR_BLACK, fmt=FMT_INT)

    # ── Tax Rate % ───────────────────────────────────────────────────────────
    _set(ws, R["TAXR"], COL_LBL, "   % of EBIT", italic=True, align_h="left")
    for i in range(n_hist):
        col = COL_H0 + i
        _set(ws, R["TAXR"], col,
             f"=IFERROR({_c(R['TAX'],col)}/{_c(R['EBIT'],col)},0)",
             color=CLR_BLACK, fmt=FMT_PCT, fill_color=HIST_BG)
    for i, v in enumerate(proj["tax_rate"]):
        _set(ws, R["TAXR"], col_p0 + i, v, color=CLR_BLUE, fmt=FMT_PCT)


def _write_cashflow_items(ws, cfg, col_p0):
    hist   = cfg["historical"]
    proj   = cfg["projection"]
    n_hist = len(hist["years"])
    n_proj = len(proj["years"])

    _section_header(ws, "CF_H", "Cash Flow Items",
                    hist["years"], proj["years"], col_p0)

    items = [
        ("DA",    "DAPCT",  "D&A",           hist["da"],    proj["da_pct"],    FMT_PCT),
        ("CAPEX", "CXPCT",  "CapEx",         hist["capex"], proj["capex_pct"], FMT_PCT1),
        ("NWC",   "NWPCT",  "Change in NWC", hist["nwc"],   proj["nwc_pct"],   FMT_PCT),
    ]
    for val_key, pct_key, label, hist_data, proj_pct, pct_fmt in items:
        _set(ws, R[val_key], COL_LBL, label, bold=True, align_h="left")
        for i, v in enumerate(hist_data):
            _set(ws, R[val_key], COL_H0 + i, v,
                 color=CLR_BLUE, fmt=FMT_INT, fill_color=HIST_BG)
        for i in range(n_proj):
            col = col_p0 + i
            _set(ws, R[val_key], col,
                 f"={_c(R['REV'],col)}*{_c(R[pct_key],col)}",
                 color=CLR_BLACK, fmt=FMT_INT)

        _set(ws, R[pct_key], COL_LBL, "   % of sales", italic=True, align_h="left")
        for i in range(n_hist):
            col = COL_H0 + i
            _set(ws, R[pct_key], col,
                 f"=IFERROR({_c(R[val_key],col)}/{_c(R['REV'],col)},0)",
                 color=CLR_BLACK, fmt=pct_fmt, fill_color=HIST_BG)
        for i, v in enumerate(proj_pct):
            _set(ws, R[pct_key], col_p0 + i, v, color=CLR_BLUE, fmt=pct_fmt)


def _write_dcf_section(ws, cfg, col_p0):
    hist   = cfg["historical"]
    proj   = cfg["projection"]
    n_hist = len(hist["years"])
    n_proj = len(proj["years"])

    # Header — numbered projection years (1, 2, 3 …)
    _set(ws, R["DCF_H"], COL_LBL, "DCF",
         bold=True, fill_color=HDR_BG, color=CLR_WHITE, align_h="left")
    for i, yr in enumerate(hist["years"]):
        _set(ws, R["DCF_H"], COL_H0 + i, str(yr),
             bold=True, fill_color=HDR_BG, color=CLR_WHITE, align_h="center")
    for i in range(n_proj):
        _set(ws, R["DCF_H"], col_p0 + i, str(i + 1),
             bold=True, fill_color=HDR_BG, color=CLR_WHITE, align_h="center")

    all_cols = list(range(COL_H0, col_p0 + n_proj))

    # ── Mirror IS rows ────────────────────────────────────────────────────────
    mirror_is = [
        ("DREV",   "REV",   "Revenue",       FMT_INT, True),
        ("DREVG",  "REVG",  "   % growth",   FMT_PCT, False),
        ("DEBIT",  "EBIT",  "EBIT",          FMT_INT, True),
        ("DEBITM", "EBITM", "   % margin",   FMT_PCT, False),
        ("DTAX",   "TAX",   "Taxes",         FMT_INT, True),
        ("DTAXR",  "TAXR",  "   % of EBIT",  FMT_PCT, False),
    ]
    for dst, src, label, fmt, is_bold in mirror_is:
        _set(ws, R[dst], COL_LBL, label, bold=is_bold,
             italic=(not is_bold), align_h="left")
        for col in all_cols:
            _set(ws, R[dst], col, f"={_c(R[src],col)}",
                 color=CLR_BLACK, fmt=fmt,
                 fill_color=HIST_BG if col < col_p0 else None)

    # ── EBIAT (projection years only) ────────────────────────────────────────
    _set(ws, R["EBIAT"], COL_LBL, "EBIAT", bold=True, align_h="left")
    for i in range(n_proj):
        col = col_p0 + i
        _set(ws, R["EBIAT"], col,
             f"={_c(R['DEBIT'],col)}-{_c(R['DTAX'],col)}",
             color=CLR_BLACK, fmt=FMT_INT)

    # ── Mirror CF rows ────────────────────────────────────────────────────────
    mirror_cf = [
        ("DDA",    "DA",    "DDAPCT",  "DAPCT",  "D&A",           FMT_PCT),
        ("DCAPEX", "CAPEX", "DCXPCT",  "CXPCT",  "CapEx",         FMT_PCT1),
        ("DNWC",   "NWC",   "DNWPCT",  "NWPCT",  "Change in NWC", FMT_PCT),
    ]
    for val_d, val_s, pct_d, pct_s, label, pct_fmt in mirror_cf:
        _set(ws, R[val_d], COL_LBL, label, bold=True, align_h="left")
        for col in all_cols:
            _set(ws, R[val_d], col, f"={_c(R[val_s],col)}",
                 color=CLR_BLACK, fmt=FMT_INT,
                 fill_color=HIST_BG if col < col_p0 else None)
        _set(ws, R[pct_d], COL_LBL, "   % of sales", italic=True, align_h="left")
        for col in all_cols:
            _set(ws, R[pct_d], col, f"={_c(R[pct_s],col)}",
                 color=CLR_BLACK, fmt=pct_fmt,
                 fill_color=HIST_BG if col < col_p0 else None)

    # ── Unlevered FCF ────────────────────────────────────────────────────────
    _set(ws, R["UFCF"], COL_LBL, "Unlevered FCF", bold=True, align_h="left")
    for i in range(n_proj):
        col = col_p0 + i
        _set(ws, R["UFCF"], col,
             f"={_c(R['EBIAT'],col)}+{_c(R['DDA'],col)}"
             f"-{_c(R['DCAPEX'],col)}-{_c(R['DNWC'],col)}",
             color=CLR_BLACK, fmt=FMT_INT, bold=True)

    # ── Present Value of FCF ─────────────────────────────────────────────────
    wacc_ref = _ca(R["WACC"], COL_H0)
    _set(ws, R["PVFCF"], COL_LBL, "Present Value of FCF", bold=True, align_h="left")
    for i in range(n_proj):
        col = col_p0 + i
        _set(ws, R["PVFCF"], col,
             f"={_c(R['UFCF'],col)}/((1+{wacc_ref})^{i+1})",
             color=CLR_BLACK, fmt=FMT_INT)


def _write_bridge(ws, cfg, col_p0, col_out):
    n_proj       = len(cfg["projection"]["years"])
    last_proj_col = col_p0 + n_proj - 1

    wacc_ref     = _ca(R["WACC"], COL_H0)
    tgr_ref      = _ca(R["TGR"],  COL_H0)
    last_fcf     = _c(R["UFCF"],  last_proj_col)
    pv_fcf_range = f"{_c(R['PVFCF'], col_p0)}:{_c(R['PVFCF'], last_proj_col)}"
    bc = col_out

    _set(ws, R["TV"],      COL_LBL, "Terminal Value",
         bold=True, align_h="left")
    _set(ws, R["TV"],      bc,
         f"={last_fcf}*(1+{tgr_ref})/({wacc_ref}-{tgr_ref})",
         color=CLR_BLACK, fmt=FMT_INT)

    _set(ws, R["PVTV"],    COL_LBL, "Present Value of Terminal Value",
         bold=True, align_h="left")
    _set(ws, R["PVTV"],    bc,
         f"={_c(R['TV'],bc)}/((1+{wacc_ref})^{n_proj})",
         color=CLR_BLACK, fmt=FMT_INT)

    _set(ws, R["EV"],      COL_LBL, "Enterprise Value",
         bold=True, align_h="left")
    _set(ws, R["EV"],      bc,
         f"=SUM({pv_fcf_range})+{_c(R['PVTV'],bc)}",
         color=CLR_BLACK, fmt=FMT_INT, bold=True)

    _set(ws, R["CASH"],    COL_LBL, "+ Cash",    bold=True, align_h="left")
    _set(ws, R["CASH"],    bc, cfg["cash"],       color=CLR_BLUE, fmt=FMT_INT)

    _set(ws, R["DEBTROW"], COL_LBL, "- Debt",    bold=True, align_h="left")
    _set(ws, R["DEBTROW"], bc, cfg["debt"],       color=CLR_BLUE, fmt=FMT_INT)

    ev_ref   = _c(R["EV"],      bc)
    cash_ref = _c(R["CASH"],    bc)
    debt_ref = _c(R["DEBTROW"], bc)
    _set(ws, R["EQV"],     COL_LBL, "Equity Value",
         bold=True, align_h="left")
    _set(ws, R["EQV"],     bc,
         f"={ev_ref}+{cash_ref}-{debt_ref}",
         color=CLR_BLACK, fmt=FMT_INT, bold=True)

    _set(ws, R["SHARES"],  COL_LBL, "/ Shares Outstanding",
         bold=True, align_h="left")
    _set(ws, R["SHARES"],  bc, cfg["shares"],    color=CLR_BLUE, fmt=FMT_INT)

    _set(ws, R["SP"],      COL_LBL, "Implied Share Price",
         bold=True, align_h="left")
    _set(ws, R["SP"],      bc,
         f"={_c(R['EQV'],bc)}/{_c(R['SHARES'],bc)}",
         color=CLR_BLACK, fmt=FMT_PRICE, bold=True, fill_color=YELLOW)


# ═══════════════════════════════════════════════════════════════════════════════
#  WORKBOOK ASSEMBLY
# ═══════════════════════════════════════════════════════════════════════════════

def build_model(cfg: dict, output_dir: str = ".") -> str:
    n_hist  = len(cfg["historical"]["years"])
    n_proj  = len(cfg["projection"]["years"])
    col_p0  = COL_H0 + n_hist          # first projection column
    col_out = col_p0 + n_proj + 1      # bridge output column (one gap after last proj)

    ticker = cfg["ticker"]
    wb = Workbook()
    ws = wb.active
    ws.title = f"{ticker} DCF"

    # Column widths
    ws.column_dimensions["A"].width = 30
    for idx in range(COL_H0, col_out + 2):
        ws.column_dimensions[get_column_letter(idx)].width = 13

    # Row heights
    for row_num in range(1, R["SP"] + 2):
        ws.row_dimensions[row_num].height = 16

    # Freeze header rows + label column
    ws.freeze_panes = "B12"

    _write_title(ws, cfg)
    _write_summary(ws, cfg, col_p0 + n_proj - 1, col_out)
    _write_assumptions(ws, cfg)
    _write_income_statement(ws, cfg, col_p0)
    _write_cashflow_items(ws, cfg, col_p0)
    _write_dcf_section(ws, cfg, col_p0)
    _write_bridge(ws, cfg, col_p0, col_out)

    output_file = str(Path(output_dir) / f"{ticker}_DCF_Model.xlsx")
    wb.save(output_file)
    print(f"✓  Saved → {output_file}")
    return output_file


# ═══════════════════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ═══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    config_path = sys.argv[1] if len(sys.argv) > 1 else "cmg_config.yaml"
    cfg = load_config(config_path)
    # Write output next to the config file
    output_dir = str(Path(config_path).parent)
    build_model(cfg, output_dir)
