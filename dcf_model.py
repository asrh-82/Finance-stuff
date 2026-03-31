#!/usr/bin/env python3
"""
DCF (Discounted Cash Flow) Valuation Model
===========================================
Generates a professional-grade, formatted Excel DCF model using Python + openpyxl.

Industry-standard color conventions:
  Blue  (#0000FF) — Hardcoded inputs / projection drivers  ← change these to stress-test
  Black (#000000) — Calculated formulas (auto-update when drivers change)

Sections:
  1. Summary Bar      – Implied price, today's price, upside/(downside)
  2. Assumptions      – WACC and terminal growth rate
  3. Income Statement – Revenue, EBIT, Taxes (historical + projected)
  4. Cash Flow Items  – D&A, CapEx, Δ NWC (historical + projected)
  5. DCF Analysis     – NOPAT, Unlevered FCF, PV of FCF
  6. Valuation Bridge – Terminal Value → Enterprise Value → Share Price

Usage:
    python dcf_model.py

Output:
    {TICKER}_DCF_Model.xlsx

Dependencies:
    pip install openpyxl
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════════════════════════
#  COMPANY INPUTS  ←  Edit this section for your analysis
# ═══════════════════════════════════════════════════════════════════════════════

TICKER        = "CMG"
COMPANY_NAME  = "Chipotle"
MODEL_DATE    = "2/18/26"
TODAY_PRICE   = 37.97          # Current market price per share ($)
SHARES        = 1_300_000      # Diluted shares outstanding ($ thousands)
CASH          = 1_050_000      # Cash & cash equivalents   ($ thousands)
DEBT          = 5_080_000      # Total debt                ($ thousands)
WACC          = 0.0755         # Weighted Average Cost of Capital
TGR           = 0.060          # Terminal Growth Rate

YEARS_HIST    = [2021, 2022, 2023, 2024, 2025]
YEARS_PROJ    = [2026, 2027, 2028, 2029, 2030]

# ── Historical Income Statement ($ thousands) — Source: Company 10-K filings ──
HIST_REV   = [7_550_000, 8_634_700,  9_871_600, 11_310_000, 11_930_000]
HIST_EBIT  = [  804_943, 1_160_403,  1_557_813,  1_916_333,  1_935_798]
HIST_TAXES = [  158_574,   282_400,    391_800,    476_100,    473_800]

# ── Historical Cash Flow Items ($ thousands) ──────────────────────────────────
HIST_DA    = [254_700, 286_800, 319_400, 335_000, 361_400]
HIST_CAPEX = [442_479, 479_164, 560_731, 593_603, 666_156]
HIST_NWC   = [805_467, 613_027, 642_900, 703_854, 740_351]

# ── Projection Drivers (2026–2030)  ←  BLUE cells — change to run scenarios ──
PROJ_REV_GRW   = [0.10,  0.10,  0.11,  0.12,  0.13]   # Revenue growth rate
PROJ_EBIT_MGN  = [0.17,  0.17,  0.18,  0.19,  0.19]   # EBIT margin (% of revenue)
PROJ_TAX_RATE  = [0.24,  0.24,  0.24,  0.24,  0.24]   # Effective tax rate
PROJ_DA_PCT    = [0.030, 0.030, 0.030, 0.030, 0.030]   # D&A % of revenue
PROJ_CAPEX_PCT = [0.057, 0.058, 0.059, 0.060, 0.061]   # CapEx % of revenue
PROJ_NWC_PCT   = [0.050, 0.050, 0.050, 0.050, 0.050]   # Δ NWC % of revenue


# ═══════════════════════════════════════════════════════════════════════════════
#  LAYOUT — Row / Column constants (change here if you need to add rows)
# ═══════════════════════════════════════════════════════════════════════════════

COL_LBL  = 1   # Column A — row labels
COL_H0   = 2   # Column B — first historical year (2021)
COL_P0   = 7   # Column G — first projection year (2026)
COL_P4   = 11  # Column K — last  projection year (2030)
COL_OUT  = 13  # Column M — single-value bridge outputs

N_HIST = len(YEARS_HIST)
N_PROJ = len(YEARS_PROJ)

R = {
    # ── Summary ──────────────────────────────────────────────────────────────
    "TITLE":    1,
    "SUM1":     3,
    "SUM2":     4,
    # ── Assumptions ──────────────────────────────────────────────────────────
    "AH":       6,
    "VALASSUMP":7,
    "WACC":     8,
    "TGR":      9,
    # ── Income Statement ─────────────────────────────────────────────────────
    "IS_H":    11,
    "REV":     12,
    "REVG":    13,
    "EBIT":    14,
    "EBITM":   15,
    "TAX":     16,
    "TAXR":    17,
    # ── Cash Flow Items ───────────────────────────────────────────────────────
    "CF_H":    19,
    "DA":      20,
    "DAPCT":   21,
    "CAPEX":   22,
    "CXPCT":   23,
    "NWC":     24,
    "NWPCT":   25,
    # ── DCF Analysis ─────────────────────────────────────────────────────────
    "DCF_H":   27,
    "DREV":    28,
    "DREVG":   29,
    "DEBIT":   30,
    "DEBITM":  31,
    "DTAX":    32,
    "DTAXR":   33,
    "PAR":     35,   # NOPAT = EBIT − Taxes
    "DDA":     36,
    "DDAPCT":  37,
    "DCAPEX":  38,
    "DCXPCT":  39,
    "DNWC":    40,
    "DNWPCT":  41,
    "UFCF":    43,   # Unlevered FCF = NOPAT + D&A − CapEx − ΔNWC
    "PVFCF":   44,   # PV of FCF     = UFCF / (1 + WACC)^n
    # ── Valuation Bridge ─────────────────────────────────────────────────────
    "TV":      46,   # Terminal Value
    "PVTV":    47,   # PV of Terminal Value
    "EV":      49,   # Enterprise Value
    "CASH":    50,
    "DEBTROW": 51,
    "EQV":     52,   # Equity Value
    "SHARES":  53,
    "SP":      54,   # Implied Share Price  ← drives SUM1
}


# ═══════════════════════════════════════════════════════════════════════════════
#  STYLE HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

# Palette
CLR_BLUE   = "0000FF"   # Hardcoded inputs
CLR_BLACK  = "000000"   # Formulas
CLR_WHITE  = "FFFFFF"
CLR_GRAY   = "808080"
HDR_BG     = "1F3864"   # Dark navy — section headers
ASSUMP_BG  = "D6E4F0"   # Light blue — assumption cells
HIST_BG    = "F2F2F2"   # Light grey — historical data
YELLOW     = "FFFF00"   # Key output highlight

# Number formats
FMT_INT    = '#,##0_);(#,##0);"-"'   # Whole-number thousands, zero → "-"
FMT_PCT    = '0.0%;(0.0%);"-"'       # Percentage, zero → "-"
FMT_PRICE  = '$#,##0.00'             # Dollar price


def _font(bold=False, color=CLR_BLACK, size=9, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _align(h="right", v="center"):
    return Alignment(horizontal=h, vertical=v)

def _c(row, col):
    """Relative cell reference, e.g. G12."""
    return f"{get_column_letter(col)}{row}"

def _ca(row, col):
    """Absolute cell reference, e.g. $G$12."""
    return f"${get_column_letter(col)}${row}"


def _set(ws, row, col, value,
         bold=False, color=CLR_BLACK, fill_color=None,
         align_h="right", fmt=None, italic=False, size=9):
    """Write a value/formula to a cell and apply formatting."""
    cell = ws.cell(row=row, column=col)
    cell.value  = value
    cell.font   = _font(bold=bold, color=color, size=size, italic=italic)
    cell.alignment = _align(align_h)
    if fill_color:
        cell.fill = _fill(fill_color)
    if fmt:
        cell.number_format = fmt


# ═══════════════════════════════════════════════════════════════════════════════
#  SECTION WRITERS
# ═══════════════════════════════════════════════════════════════════════════════

def _write_title(ws):
    _set(ws, R["TITLE"], COL_LBL,
         f"{COMPANY_NAME} — DCF Valuation Model",
         bold=True, size=13, align_h="left")


def _write_summary(ws):
    r1, r2 = R["SUM1"], R["SUM2"]
    sp_ref      = _ca(R["SP"],  COL_OUT)   # Implied share price cell
    implied_ref = _ca(r1, COL_H0 + 3)      # $E$3
    today_ref   = _ca(r1, COL_H0 + 6)      # $H$3

    # Row 1
    _set(ws, r1, COL_LBL,       "Ticker:",              bold=True, align_h="left")
    _set(ws, r1, COL_H0,        TICKER,                 bold=True, color=CLR_BLUE,
         fill_color=YELLOW, align_h="center")
    _set(ws, r1, COL_H0 + 2,   "Implied Share Price:",  bold=True, align_h="right")
    _set(ws, r1, COL_H0 + 3,   f"={sp_ref}",
         color=CLR_BLACK, fmt=FMT_PRICE, align_h="center")
    _set(ws, r1, COL_H0 + 5,   "Today's Share Price:",  bold=True, align_h="right")
    _set(ws, r1, COL_H0 + 6,   TODAY_PRICE,
         color=CLR_BLUE, fmt=FMT_PRICE, align_h="center")
    _set(ws, r1, COL_H0 + 7,   "Upside / (Downside):", bold=True, align_h="right")
    _set(ws, r1, COL_H0 + 8,
         f"=({implied_ref}-{today_ref})/{today_ref}",
         color=CLR_BLACK, fmt=FMT_PCT, align_h="center")

    # Row 2
    _set(ws, r2, COL_LBL, "Date:",      bold=True, align_h="left")
    _set(ws, r2, COL_H0,  MODEL_DATE,   color=CLR_BLUE, align_h="center")
    _set(ws, r2, COL_P4,  "All numbers in thousands",
         italic=True, size=8, align_h="right", color=CLR_GRAY)


def _write_assumptions(ws):
    _set(ws, R["AH"],       COL_LBL, "Assumptions",
         bold=True, fill_color=HDR_BG, color=CLR_WHITE, align_h="left", size=10)
    _set(ws, R["VALASSUMP"],COL_LBL, "Valuation Assumptions",
         bold=True, align_h="left")
    _set(ws, R["WACC"],     COL_LBL, "WACC",
         bold=True, align_h="left")
    _set(ws, R["WACC"],     COL_H0,  WACC,
         color=CLR_BLUE, fmt=FMT_PCT, fill_color=ASSUMP_BG, align_h="center")
    _set(ws, R["TGR"],      COL_LBL, "TGR",
         bold=True, align_h="left")
    _set(ws, R["TGR"],      COL_H0,  TGR,
         color=CLR_BLUE, fmt=FMT_PCT, fill_color=ASSUMP_BG, align_h="center")


def _section_header(ws, row_key, label):
    """Write a dark-navy section header with all year labels."""
    _set(ws, R[row_key], COL_LBL, label,
         bold=True, fill_color=HDR_BG, color=CLR_WHITE, align_h="left")
    for i, yr in enumerate(YEARS_HIST):
        _set(ws, R[row_key], COL_H0 + i, str(yr),
             bold=True, fill_color=HDR_BG, color=CLR_WHITE, align_h="center")
    for i, yr in enumerate(YEARS_PROJ):
        _set(ws, R[row_key], COL_P0 + i, str(yr),
             bold=True, fill_color=HDR_BG, color=CLR_WHITE, align_h="center")


def _write_income_statement(ws):
    _section_header(ws, "IS_H", "Income Statement")

    # ── Revenue ──────────────────────────────────────────────────────────────
    _set(ws, R["REV"], COL_LBL, "Revenue", bold=True, align_h="left")
    for i, v in enumerate(HIST_REV):
        _set(ws, R["REV"], COL_H0 + i, v,
             color=CLR_BLUE, fmt=FMT_INT, fill_color=HIST_BG)
    for i in range(N_PROJ):
        col  = COL_P0 + i
        prev = _c(R["REV"],  col - 1)
        grw  = _c(R["REVG"], col)
        _set(ws, R["REV"], col, f"={prev}*(1+{grw})", color=CLR_BLACK, fmt=FMT_INT)

    # ── Revenue Growth % ─────────────────────────────────────────────────────
    _set(ws, R["REVG"], COL_LBL, "   % growth", italic=True, align_h="left")
    for i in range(1, N_HIST):
        col = COL_H0 + i
        _set(ws, R["REVG"], col,
             f"=IFERROR(({_c(R['REV'],col)}-{_c(R['REV'],col-1)})/{_c(R['REV'],col-1)},0)",
             color=CLR_BLACK, fmt=FMT_PCT, fill_color=HIST_BG)
    for i, v in enumerate(PROJ_REV_GRW):
        _set(ws, R["REVG"], COL_P0 + i, v, color=CLR_BLUE, fmt=FMT_PCT)

    # ── EBIT ─────────────────────────────────────────────────────────────────
    _set(ws, R["EBIT"], COL_LBL, "EBIT", bold=True, align_h="left")
    for i, v in enumerate(HIST_EBIT):
        _set(ws, R["EBIT"], COL_H0 + i, v,
             color=CLR_BLUE, fmt=FMT_INT, fill_color=HIST_BG)
    for i in range(N_PROJ):
        col = COL_P0 + i
        _set(ws, R["EBIT"], col,
             f"={_c(R['REV'],col)}*{_c(R['EBITM'],col)}",
             color=CLR_BLACK, fmt=FMT_INT)

    # ── EBIT Margin % ────────────────────────────────────────────────────────
    _set(ws, R["EBITM"], COL_LBL, "   % of sales", italic=True, align_h="left")
    for i in range(N_HIST):
        col = COL_H0 + i
        _set(ws, R["EBITM"], col,
             f"=IFERROR({_c(R['EBIT'],col)}/{_c(R['REV'],col)},0)",
             color=CLR_BLACK, fmt=FMT_PCT, fill_color=HIST_BG)
    for i, v in enumerate(PROJ_EBIT_MGN):
        _set(ws, R["EBITM"], COL_P0 + i, v, color=CLR_BLUE, fmt=FMT_PCT)

    # ── Taxes ────────────────────────────────────────────────────────────────
    _set(ws, R["TAX"], COL_LBL, "Taxes", bold=True, align_h="left")
    for i, v in enumerate(HIST_TAXES):
        _set(ws, R["TAX"], COL_H0 + i, v,
             color=CLR_BLUE, fmt=FMT_INT, fill_color=HIST_BG)
    for i in range(N_PROJ):
        col = COL_P0 + i
        _set(ws, R["TAX"], col,
             f"={_c(R['EBIT'],col)}*{_c(R['TAXR'],col)}",
             color=CLR_BLACK, fmt=FMT_INT)

    # ── Tax Rate % ───────────────────────────────────────────────────────────
    _set(ws, R["TAXR"], COL_LBL, "   % of EBIT", italic=True, align_h="left")
    for i in range(N_HIST):
        col = COL_H0 + i
        _set(ws, R["TAXR"], col,
             f"=IFERROR({_c(R['TAX'],col)}/{_c(R['EBIT'],col)},0)",
             color=CLR_BLACK, fmt=FMT_PCT, fill_color=HIST_BG)
    for i, v in enumerate(PROJ_TAX_RATE):
        _set(ws, R["TAXR"], COL_P0 + i, v, color=CLR_BLUE, fmt=FMT_PCT)


def _write_cashflow_items(ws):
    _section_header(ws, "CF_H", "Cash Flow Items")

    items = [
        ("DA",    "DAPCT",  "D&A",          HIST_DA,    PROJ_DA_PCT),
        ("CAPEX", "CXPCT",  "CapEx",        HIST_CAPEX, PROJ_CAPEX_PCT),
        ("NWC",   "NWPCT",  "Change in NWC",HIST_NWC,   PROJ_NWC_PCT),
    ]
    for val_key, pct_key, label, hist_data, proj_pct in items:
        _set(ws, R[val_key], COL_LBL, label, bold=True, align_h="left")
        for i, v in enumerate(hist_data):
            _set(ws, R[val_key], COL_H0 + i, v,
                 color=CLR_BLUE, fmt=FMT_INT, fill_color=HIST_BG)
        for i in range(N_PROJ):
            col = COL_P0 + i
            _set(ws, R[val_key], col,
                 f"={_c(R['REV'],col)}*{_c(R[pct_key],col)}",
                 color=CLR_BLACK, fmt=FMT_INT)

        _set(ws, R[pct_key], COL_LBL, "   % of sales", italic=True, align_h="left")
        for i in range(N_HIST):
            col = COL_H0 + i
            _set(ws, R[pct_key], col,
                 f"=IFERROR({_c(R[val_key],col)}/{_c(R['REV'],col)},0)",
                 color=CLR_BLACK, fmt=FMT_PCT, fill_color=HIST_BG)
        for i, v in enumerate(proj_pct):
            _set(ws, R[pct_key], COL_P0 + i, v, color=CLR_BLUE, fmt=FMT_PCT)


def _write_dcf_section(ws):
    # Header with year-number labels for projection period
    _set(ws, R["DCF_H"], COL_LBL, "DCF",
         bold=True, fill_color=HDR_BG, color=CLR_WHITE, align_h="left")
    for i, yr in enumerate(YEARS_HIST):
        _set(ws, R["DCF_H"], COL_H0 + i, str(yr),
             bold=True, fill_color=HDR_BG, color=CLR_WHITE, align_h="center")
    for i in range(N_PROJ):
        _set(ws, R["DCF_H"], COL_P0 + i, str(i + 1),   # Year 1 … 5
             bold=True, fill_color=HDR_BG, color=CLR_WHITE, align_h="center")

    all_cols = list(range(COL_H0, COL_P0 + N_PROJ))

    # Mirror IS rows
    mirror_is = [
        ("DREV",   "REV",   "Revenue",       FMT_INT, True),
        ("DREVG",  "REVG",  "   % growth",   FMT_PCT, False),
        ("DEBIT",  "EBIT",  "EBIT",          FMT_INT, True),
        ("DEBITM", "EBITM", "   % margin",   FMT_PCT, False),
        ("DTAX",   "TAX",   "Taxes",         FMT_INT, True),
        ("DTAXR",  "TAXR",  "   % of EBIT",  FMT_PCT, False),
    ]
    for dst, src, label, fmt, is_bold in mirror_is:
        _set(ws, R[dst], COL_LBL, label, bold=is_bold,
             italic=(not is_bold), align_h="left")
        for col in all_cols:
            _set(ws, R[dst], col, f"={_c(R[src],col)}",
                 color=CLR_BLACK, fmt=fmt,
                 fill_color=HIST_BG if col < COL_P0 else None)

    # ── NOPAT (PAR = EBIT − Taxes) — projection years only ──────────────────
    _set(ws, R["PAR"], COL_LBL, "PAR  (NOPAT)", bold=True, align_h="left")
    for i in range(N_PROJ):
        col = COL_P0 + i
        _set(ws, R["PAR"], col,
             f"={_c(R['DEBIT'],col)}-{_c(R['DTAX'],col)}",
             color=CLR_BLACK, fmt=FMT_INT)

    # Mirror CF rows
    mirror_cf = [
        ("DDA",    "DA",    "DDAPCT",  "DAPCT",  "D&A",          PROJ_DA_PCT),
        ("DCAPEX", "CAPEX", "DCXPCT",  "CXPCT",  "CapEx",        PROJ_CAPEX_PCT),
        ("DNWC",   "NWC",   "DNWPCT",  "NWPCT",  "Change in NWC",PROJ_NWC_PCT),
    ]
    for val_d, val_s, pct_d, pct_s, label, _ in mirror_cf:
        _set(ws, R[val_d], COL_LBL, label, bold=True, align_h="left")
        for col in all_cols:
            _set(ws, R[val_d], col, f"={_c(R[val_s],col)}",
                 color=CLR_BLACK, fmt=FMT_INT,
                 fill_color=HIST_BG if col < COL_P0 else None)
        _set(ws, R[pct_d], COL_LBL, "   % of sales", italic=True, align_h="left")
        for col in all_cols:
            _set(ws, R[pct_d], col, f"={_c(R[pct_s],col)}",
                 color=CLR_BLACK, fmt=FMT_PCT,
                 fill_color=HIST_BG if col < COL_P0 else None)

    # ── Unlevered FCF ────────────────────────────────────────────────────────
    _set(ws, R["UFCF"], COL_LBL, "Unlevered FCF", bold=True, align_h="left")
    for i in range(N_PROJ):
        col   = COL_P0 + i
        par   = _c(R["PAR"],    col)
        da    = _c(R["DDA"],    col)
        capex = _c(R["DCAPEX"], col)
        nwc   = _c(R["DNWC"],   col)
        _set(ws, R["UFCF"], col,
             f"={par}+{da}-{capex}-{nwc}",
             color=CLR_BLACK, fmt=FMT_INT, bold=True)

    # ── Present Value of FCF ─────────────────────────────────────────────────
    wacc_ref = _ca(R["WACC"], COL_H0)   # $B$8
    _set(ws, R["PVFCF"], COL_LBL, "Present Value of FCF", bold=True, align_h="left")
    for i in range(N_PROJ):
        col = COL_P0 + i
        n   = i + 1
        _set(ws, R["PVFCF"], col,
             f"={_c(R['UFCF'],col)}/((1+{wacc_ref})^{n})",
             color=CLR_BLACK, fmt=FMT_INT)


def _write_bridge(ws):
    """Terminal Value → Enterprise Value → Share Price."""
    wacc_ref     = _ca(R["WACC"], COL_H0)          # $B$8
    tgr_ref      = _ca(R["TGR"],  COL_H0)          # $B$9
    last_fcf     = _c(R["UFCF"],  COL_P4)          # K43
    pv_fcf_range = (f"{_c(R['PVFCF'], COL_P0)}:"
                    f"{_c(R['PVFCF'], COL_P4)}")   # G44:K44
    bc = COL_OUT  # Column M

    # Terminal Value
    _set(ws, R["TV"],  COL_LBL, "Terminal Value",              bold=True, align_h="left")
    _set(ws, R["TV"],  bc,
         f"={last_fcf}*(1+{tgr_ref})/({wacc_ref}-{tgr_ref})",
         color=CLR_BLACK, fmt=FMT_INT)

    # PV of Terminal Value
    _set(ws, R["PVTV"], COL_LBL, "Present Value of Terminal Value", bold=True, align_h="left")
    _set(ws, R["PVTV"], bc,
         f"={_c(R['TV'],bc)}/((1+{wacc_ref})^{N_PROJ})",
         color=CLR_BLACK, fmt=FMT_INT)

    # Enterprise Value
    _set(ws, R["EV"],  COL_LBL, "Enterprise Value",            bold=True, align_h="left")
    _set(ws, R["EV"],  bc,
         f"=SUM({pv_fcf_range})+{_c(R['PVTV'],bc)}",
         color=CLR_BLACK, fmt=FMT_INT, bold=True)

    # + Cash
    _set(ws, R["CASH"],    COL_LBL, "+ Cash",                  bold=True, align_h="left")
    _set(ws, R["CASH"],    bc, CASH,   color=CLR_BLUE, fmt=FMT_INT)

    # − Debt
    _set(ws, R["DEBTROW"], COL_LBL, "- Debt",                  bold=True, align_h="left")
    _set(ws, R["DEBTROW"], bc, DEBT,   color=CLR_BLUE, fmt=FMT_INT)

    # Equity Value
    ev_ref   = _c(R["EV"],      bc)
    cash_ref = _c(R["CASH"],    bc)
    debt_ref = _c(R["DEBTROW"], bc)
    _set(ws, R["EQV"], COL_LBL, "Equity Value",                bold=True, align_h="left")
    _set(ws, R["EQV"], bc,
         f"={ev_ref}+{cash_ref}-{debt_ref}",
         color=CLR_BLACK, fmt=FMT_INT, bold=True)

    # ÷ Shares
    _set(ws, R["SHARES"], COL_LBL, "/ Shares Outstanding",     bold=True, align_h="left")
    _set(ws, R["SHARES"], bc, SHARES, color=CLR_BLUE, fmt=FMT_INT)

    # Implied Share Price
    _set(ws, R["SP"], COL_LBL, "Implied Share Price",        bold=True, align_h="left")
    _set(ws, R["SP"], bc,
         f"={_c(R['EQV'],bc)}/{_c(R['SHARES'],bc)}",
         color=CLR_BLACK, fmt=FMT_PRICE, bold=True, fill_color=YELLOW)


# ═══════════════════════════════════════════════════════════════════════════════
#  WORKBOOK ASSEMBLY
# ═══════════════════════════════════════════════════════════════════════════════

def build_model() -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = f"{TICKER} DCF"

    # Column widths
    ws.column_dimensions["A"].width = 30
    for idx in range(COL_H0, COL_OUT + 2):
        ws.column_dimensions[get_column_letter(idx)].width = 13

    # Row heights (slightly taller for readability)
    for row_num in range(1, R["SP"] + 2):
        ws.row_dimensions[row_num].height = 16

    # Freeze: keep label column + header rows in view while scrolling
    ws.freeze_panes = "B12"

    # Write all sections
    _write_title(ws)
    _write_summary(ws)
    _write_assumptions(ws)
    _write_income_statement(ws)
    _write_cashflow_items(ws)
    _write_dcf_section(ws)
    _write_bridge(ws)

    output_file = f"{TICKER}_DCF_Model.xlsx"
    wb.save(output_file)
    print(f"✓  Model saved → {output_file}")
    return output_file


if __name__ == "__main__":
    build_model()
